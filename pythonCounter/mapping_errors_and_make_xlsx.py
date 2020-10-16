import pymysql
from mapping_using_regex import error_to_solve
from mapping_wrong_answer import check_output, check_source_code, check_source_code_using_json
from openpyxl import load_workbook, Workbook

load_wb = load_workbook("/home/oem/Desktop/labeled_error9.xlsx", data_only=True)
load_ws = load_wb['Sheet']
write_wb = Workbook()
write_ws = write_wb.active

write_ws.append(['index', 'error-type', 'error-msg', 'error-classification', 'wrong algorithm-classification', 'detailed'])

db = pymysql.connect(host='192.168.56.103', port=3306, user='root', password='1234', db='domjudge', charset='utf8')
cursor = db.cursor()

#3471, 5470
# sql = '''
#
# select s.submitid, l.name, j.result from `submission` as s join `language` as l join `judging`
# as j on s.submitid=j.submitid and s.langid = l.langid where s.submitid>=3471 and s.submitid<=5470 order by s.submitid
#
# '''
sql = '''

select s.submitid, l.name, j.result from `submission` as s join `language` as l join `judging`
as j on s.submitid=j.submitid and s.langid = l.langid where s.submitid>=3471 and s.submitid<=5470
and j.result!="correct" order by s.submitid

'''
cursor.execute(sql)
submission_data = cursor.fetchall()
# 0 : submitid, 1 : lang, 2 : error-type

check_wrong_output = 0

cur_index_in_excel = 2
submission_data_n = len(submission_data)

for i in range(0, submission_data_n) :
# for submit in submission_data :
    submit = submission_data[i]
    submit_id = submit[0]

    if i+1<submission_data_n and submit_id == submission_data[i+1][0] :
        continue

    lang = submit[1]
    error_type = submit[2]
    print(load_ws['A'+str(cur_index_in_excel)].value )

    while load_ws['A' + str(cur_index_in_excel)].value is None:
        cur_index_in_excel = cur_index_in_excel+1

    if submit_id != load_ws['A'+str(cur_index_in_excel)].value :
        continue

    dict_to_new_excel = [load_ws['A'+str(cur_index_in_excel)].value, load_ws['B'+str(cur_index_in_excel)].value, load_ws['C'+str(cur_index_in_excel)].value]
    dict_to_new_excel = [submit_id, error_type]

    if submit_id == 3489:
        print("hi")

    print("===================================")
    print(submit_id)
    print(error_type+"\n")
    if error_type == 'run-error':
        sql = '''
                select j.submitid, jr.output_error from judging as j join judging_run as jr
                on j.judgingid = jr.judgingid where j.submitid=%s and jr.runresult="run-error";
                '''
        cursor.execute(sql, submit_id)
        run_data = cursor.fetchall()

        print(run_data[0][1].decode())
        detailed_error = run_data[0][1].decode().strip().split("\n")
        run_error_msg = ""

        if lang == "Java":
            run_error_msg = detailed_error[0]
        elif lang == "Python 3" or lang == "Python 2":
            run_error_msg = detailed_error[-1]
        else :
            if len(detailed_error) <= 1:
                print("check this submission")
                write_ws.append(dict_to_new_excel)
                cur_index_in_excel = cur_index_in_excel + 1
                continue
            run_error_msg = detailed_error[1]

        solve = error_to_solve(run_error_msg)
        check_wrong_output = check_wrong_output + 1
        print(solve)
        dict_to_new_excel.append(solve)

    elif error_type == 'compiler-error':
        sql = '''
                select submitid, output_compile from judging where submitid=%s;
                '''

        cursor.execute(sql, submit_id)
        compile_data = cursor.fetchall()

        print(compile_data[0][1].decode())
        detailed_error = compile_data[0][1].decode().strip().split("\n")

        compile_error_msg = ""
        if lang == "Java":
            compile_error_msg = detailed_error[1]
        else :
            compile_error_msg = detailed_error[-1]
        solve = error_to_solve(compile_error_msg)
        check_wrong_output = check_wrong_output + 1
        print(solve)
        dict_to_new_excel.append(solve)

    elif error_type == 'wrong-answer' or error_type == 'correct':
        sql = '''
            select j.submitid, t.testcaseid, jr.runresult, jr.output_run, t.output
            from `judging` as j join `judging_run` as jr on j.judgingid = jr.judgingid
            join `testcase` as t on jr.testcaseid=t.testcaseid
            where j.submitid=%s and jr.runresult!="timelimit"
        '''

        cursor.execute(sql, submit_id)
        wrong_answer_data = cursor.fetchall()
        solve = check_output(wrong_answer_data)

        if solve :
            print(solve)

        else :
            sql = '''
                select w.submitid, c.submitid, w.probid from submission as w join submission as c on w.teamid=c.teamid
                join judging as j on c.submitid=j.submitid where w.submitid=%s and w.probid=c.probid
                and j.result="correct" and w.langid=c.langid and w.submitid<=c.submitid limit 1
            '''
            cursor.execute(sql, submit_id)
            submit_data = cursor.fetchall()

            if len(submit_data) > 0 :

                sql = '''
                    select sourcecode from submission_file where submitid=%s
                '''
                cursor.execute(sql, submit_id)
                wrong_source_code = cursor.fetchall()

                if submit_data[0][2] < 32 or submit_data[0][2] > 56:
                    write_ws.append(dict_to_new_excel)
                    cur_index_in_excel = cur_index_in_excel + 1
                    continue

                # correct_submit_id = submit_data[0][1]
                # cursor.execute(sql, correct_submit_id)
                # correct_source_code = cursor.fetchall()
                #
                # solve = check_source_code(wrong_source_code[0][0].decode(), correct_source_code[0][0].decode(), lang)

                sorted_dict, solve, similarity = check_source_code_using_json(wrong_source_code[0][0].decode(), lang, str(submit_data[0][2]))
                #print("submitid : ", submit_id, " probid : ", submit_data[0][2], "similarity : ", similarity)
                if sorted_dict != None :
                    print("정답 소스코드 평균과의 차이 : \n", sorted_dict)
                print(solve)
                print("============================================================")
        dict_to_new_excel.append(load_ws['D'+str(cur_index_in_excel)].value)
        dict_to_new_excel.append(load_ws['E' + str(cur_index_in_excel)].value)
        dict_to_new_excel.append(load_ws['F' + str(cur_index_in_excel)].value)
        dict_to_new_excel.append(load_ws['G' + str(cur_index_in_excel)].value)
        dict_to_new_excel.append(str(solve))
        # dict_to_new_excel.append(similarity)

    else :
        print(error_type)

    write_ws.append(dict_to_new_excel)
    cur_index_in_excel = cur_index_in_excel + 1

# load_wb.close()
write_wb.save("/home/oem/Desktop/labeled_error10.xlsx")
db.close()

# source_code = '''
# import java.io.BufferedReader;
# import java.io.IOException;
# import java.io.InputStreamReader;
# import java.util.Collections;
# import java.util.HashMap;
# import java.util.Iterator;
# import java.util.LinkedList;
# import java.util.PriorityQueue;
# import java.util.StringTokenizer;
# import java.util.TreeSet;
#
# class Edge implements Comparable<Edge> {
# 	String s, e;
# 	int cost;
# 	public Edge(String s, String e, int cost) {
# 		this.s = s;
# 		this.e = e;
# 		this.cost = cost;
# 	}
# 	@Override
# 	public int compareTo(Edge o) {
# 		return Integer.compare(this.cost, o.cost);
# 	}
# 	@Override
# 	public String toString() {
# 		return Integer.toString(cost);
# 	}
#
# }
#
# public class P4_MST {
# 	static HashMap<String, Boolean> visited;
# 	static HashMap<String, String> tree;
# 	static int totalCost = 0;
# 	static int totalCost2 = 0;
# 	public static String find(String x) {
# 		if(x.equals(tree.getOrDefault(x, x))) {
# 			return x;
# 		}
# 		tree.put(x, find(tree.getOrDefault(x, x)));
# 		return tree.get(x);
# 	}
#
# 	public static void union(String x, String y) {
# 		String xRoot = find(x);
# 		tree.put(xRoot, y);
# 	}
#
# 	public static void main(String[] args) throws IOException {
# 		BufferedReader br = new BufferedReader(new InputStreamReader(System.in));
# 		StringTokenizer st = new StringTokenizer(br.readLine());
# 		int n = Integer.parseInt(st.nextToken()); //node
# 		int m = Integer.parseInt(st.nextToken()); //edge
# 		tree = new HashMap<String, String>();
# 		visited = new HashMap<String, Boolean>();
# 		//input nodes
# 		st = new StringTokenizer(br.readLine());
# 		while(st.hasMoreTokens()) {
# 			String token = st.nextToken();
# 		}
#
# 		LinkedList<Edge> list = new LinkedList<Edge>();
# 		for(int i = 0; i < m; i++) {
# 			st = new StringTokenizer(br.readLine());
# 			String s = st.nextToken();
# 			String e = st.nextToken();
# 			int cost = Integer.parseInt(st.nextToken());
# 			list.add(new Edge(s, e, cost));
# 		}
#
# 		PriorityQueue<Edge> used = new PriorityQueue<Edge>();
# 		//Make 1st MST
# 		Collections.sort(list);
# 		Iterator<Edge> iter = list.iterator();
# 		while(iter.hasNext()) {
# 			Edge edge = iter.next();
# 			String x = edge.s;
# 			String y = edge.e;
# 			String xRoot = find(x);
# 			String yRoot = find(y);
# 			if(xRoot.equals(yRoot)) {
# 				continue;
# 			}
# 			used.add(edge);
# 			union(x, y);
# 			totalCost += edge.cost;
# 		}
#
# 		//2st MST
# 		int ans = Integer.MAX_VALUE;
# 		while(!used.isEmpty()) {
# 			totalCost2 = 0;
# 			tree = new HashMap<String, String>();
# 			//MST에 사용된 간선 중 하나를 지운다.
# 			Edge remove = used.poll();
# 			LinkedList<Edge> list_ = new LinkedList<Edge>();
# 			list_.addAll(list);
# 			list_.remove(remove);
#
# 			//MST 생성
# 			Edge edge = list_.poll();
# 			String x = edge.s;
# 			String y = edge.e;
# 			union(x, y);
# 			totalCost2 += edge.cost;
# 			iter = list_.iterator();
# 			while(iter.hasNext()) {
# 				edge = iter.next();
# 				x = edge.s;
# 				y = edge.e;
# 				String xRoot = find(x);
# 				String yRoot = find(y);
# 				if(xRoot.equals(yRoot)) {
# 					continue;
# 				}
# 				union(x, y);
# 				totalCost2 += edge.cost;
# 			}
#
# 			//tree가 모두 연결되어있는가?
# 			int uniqueValues = new TreeSet<String>(tree.values()).size();
# 			if(uniqueValues > 1) {
# 				continue;
# 			}
# 			ans = Math.min(totalCost2, ans);
# 		}
#
#
# 		System.out.println(ans);
# 		br.close();
# 	}
# }
#
# '''
# solve = check_source_code_using_json(source_code, "Java", "34")
# print(solve)


